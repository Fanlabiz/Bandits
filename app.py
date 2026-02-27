"""Application Streamlit pour placement de bâtiments - Optimisation maximale des boosts"""
import streamlit as st
import pandas as pd
import numpy as np
from dataclasses import dataclass
from typing import List, Tuple, Optional
import copy
import io
import plotly.graph_objects as go
import plotly.express as px
import sys
from collections import defaultdict
import math
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

print("=== DÉMARRAGE DE L'APPLICATION OPTIMISATION BOOSTS ===")

@dataclass
class Building:
    name: str
    length: int
    width: int
    quantity: int
    culture_produced: float
    radius: int
    boost_25: float
    boost_50: float
    boost_100: float
    
    def get_dimensions(self, orientation: str) -> Tuple[int, int]:
        if orientation == 'H':
            return self.length, self.width
        else:
            return self.width, self.length
    
    def is_producer(self) -> bool:
        return self.culture_produced > 0 and self.radius > 0
    
    def can_be_boosted(self) -> bool:
        return (self.boost_25 > 0 or self.boost_50 > 0 or self.boost_100 > 0)
    
    def get_area(self) -> int:
        return self.length * self.width

class BoostOptimizer:
    """Classe dédiée à l'optimisation des boosts"""
    
    def __init__(self, producers: List[Building], consumers: List[Building]):
        self.producers = producers
        self.consumers = consumers
        self.boost_matrix = {}  # (producer_id, consumer_id) -> potentiel de boost
        
    def calculate_boost_potential(self, producer: Building, consumer: Building) -> float:
        """Calcule le potentiel de boost d'un producteur sur un consommateur"""
        if not consumer.can_be_boosted():
            return 0
        
        # Plus la culture du producteur est élevée, mieux c'est
        # Plus les seuils du consommateur sont bas, mieux c'est
        min_threshold = min(t for t in [consumer.boost_25, consumer.boost_50, consumer.boost_100] if t > 0)
        return producer.culture_produced / min_threshold * 100
    
    def find_best_clusters(self) -> List[List[Building]]:
        """Identifie les clusters de bâtiments qui se boostent mutuellement"""
        clusters = []
        remaining_consumers = self.consumers.copy()
        
        # Pour chaque producteur, créer un cluster avec les consommateurs compatibles
        for producer in self.producers:
            cluster = [producer]
            compatible = []
            
            for consumer in remaining_consumers:
                potential = self.calculate_boost_potential(producer, consumer)
                if potential > 10:  # Seuil de compatibilité
                    compatible.append(consumer)
            
            # Trier par potentiel de boost
            compatible.sort(key=lambda c: -self.calculate_boost_potential(producer, c))
            
            # Prendre les meilleurs jusqu'à saturation du rayon
            cluster.extend(compatible[:5])  # Max 5 consommateurs par producteur
            
            # Retirer les consommateurs placés
            for c in compatible[:5]:
                if c in remaining_consumers:
                    remaining_consumers.remove(c)
            
            if len(cluster) > 1:
                clusters.append(cluster)
        
        return clusters

class BuildingPlacer:
    def __init__(self, terrain_grid: np.ndarray, buildings: List[Building]):
        self.terrain_grid = terrain_grid
        self.buildings = buildings
        self.placement_grid = None
        self.placed_buildings = []
        self.available_zones = []
        self.producer_positions = []  # (x, y, building, id, radius)
        self.consumer_positions = []  # (x, y, building, id)
        self.initialize_grids()
        
    def initialize_grids(self):
        self.placement_grid = np.zeros_like(self.terrain_grid)
        self.placement_grid[self.terrain_grid == 0] = -1
        
    def find_available_zones(self):
        visited = np.zeros_like(self.placement_grid, dtype=bool)
        self.available_zones = []
        
        for i in range(self.placement_grid.shape[0]):
            for j in range(self.placement_grid.shape[1]):
                if self.placement_grid[i, j] == 0 and not visited[i, j]:
                    zone = []
                    stack = [(i, j)]
                    visited[i, j] = True
                    
                    while stack:
                        x, y = stack.pop()
                        zone.append((x, y))
                        for dx, dy in [(0,1), (1,0), (0,-1), (-1,0)]:
                            nx, ny = x+dx, y+dy
                            if (0 <= nx < self.placement_grid.shape[0] and 
                                0 <= ny < self.placement_grid.shape[1] and
                                self.placement_grid[nx, ny] == 0 and 
                                not visited[nx, ny]):
                                visited[nx, ny] = True
                                stack.append((nx, ny))
                    
                    if zone:
                        self.available_zones.append(zone)
        
        return len(self.available_zones)
    
    def can_place(self, x: int, y: int, length: int, width: int) -> bool:
        """Vérifie si on peut placer un bâtiment à une position"""
        if x + length > self.placement_grid.shape[0] or y + width > self.placement_grid.shape[1]:
            return False
        
        for i in range(length):
            for j in range(width):
                if self.placement_grid[x + i, y + j] != 0:
                    return False
        return True
    
    def calculate_culture_at_position(self, x: int, y: int, length: int, width: int) -> float:
        """Calcule la culture reçue à une position donnée"""
        total_culture = 0
        center_x = x + length // 2
        center_y = y + width // 2
        
        for px, py, p_building, p_id, p_radius in self.producer_positions:
            dist = abs(center_x - px) + abs(center_y - py)
            if dist <= p_radius:
                total_culture += p_building.culture_produced
        
        return total_culture
    
    def find_positions_in_radius(self, producer_x: int, producer_y: int, 
                                producer_radius: int, zone: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
        """Trouve toutes les positions dans le rayon d'un producteur"""
        positions = []
        for (zx, zy) in zone:
            dist = abs(zx - producer_x) + abs(zy - producer_y)
            if dist <= producer_radius:
                positions.append((zx, zy))
        return positions
    
    def place_cluster(self, cluster: List[Building], zone: List[Tuple[int, int]]) -> bool:
        """Place un cluster de bâtiments (producteur + consommateurs) dans une zone"""
        if not zone:
            return False
        
        # Le premier élément est le producteur
        producer = cluster[0]
        consumers = cluster[1:]
        
        # Trouver une position pour le producteur
        xs = [p[0] for p in zone]
        ys = [p[1] for p in zone]
        min_x, max_x = min(xs), max(xs)
        min_y, max_y = min(ys), max(ys)
        center_x = (min_x + max_x) // 2
        center_y = (min_y + max_y) // 2
        
        # Placer le producteur
        producer_placed = False
        for orientation in ['H', 'V']:
            length, width = producer.get_dimensions(orientation)
            
            for dx in [-3, -2, -1, 0, 1, 2, 3]:
                for dy in [-3, -2, -1, 0, 1, 2, 3]:
                    x = max(min_x, min(center_x + dx, max_x - length + 1))
                    y = max(min_y, min(center_y + dy, max_y - width + 1))
                    
                    if self.can_place(x, y, length, width):
                        if self.place_building(producer, x, y, orientation):
                            producer_x = x + length // 2
                            producer_y = y + width // 2
                            producer_radius = producer.radius
                            producer_placed = True
                            break
                    if producer_placed:
                        break
                if producer_placed:
                    break
            if producer_placed:
                break
        
        if not producer_placed:
            return False
        
        # Mettre à jour la zone
        new_zone = []
        for (zx, zy) in zone:
            if self.placement_grid[zx, zy] == 0:
                new_zone.append((zx, zy))
        zone = new_zone
        
        if not zone:
            return True
        
        # Placer les consommateurs dans le rayon du producteur
        radius_positions = self.find_positions_in_radius(producer_x, producer_y, producer_radius, zone)
        
        for consumer in consumers:
            if not radius_positions:
                break
            
            best_placement = None
            best_culture = -1
            
            for orientation in ['H', 'V']:
                length, width = consumer.get_dimensions(orientation)
                
                for (x, y) in radius_positions:
                    if x + length <= self.placement_grid.shape[0] and y + width <= self.placement_grid.shape[1]:
                        if self.can_place(x, y, length, width):
                            culture = self.calculate_culture_at_position(x, y, length, width)
                            if culture > best_culture:
                                best_culture = culture
                                best_placement = (x, y, orientation)
            
            if best_placement:
                x, y, orientation = best_placement
                self.place_building(consumer, x, y, orientation)
                
                # Mettre à jour la zone et les positions dans le rayon
                new_zone = []
                for (zx, zy) in zone:
                    if self.placement_grid[zx, zy] == 0:
                        new_zone.append((zx, zy))
                zone = new_zone
                radius_positions = self.find_positions_in_radius(producer_x, producer_y, producer_radius, zone)
        
        return True
    
    def place_building(self, building: Building, x: int, y: int, orientation: str) -> bool:
        length, width = building.get_dimensions(orientation)
        
        building_id = len(self.placed_buildings) + 1
        for i in range(length):
            for j in range(width):
                self.placement_grid[x + i, y + j] = building_id
        
        # Calculer la culture reçue
        culture_recue = self.calculate_culture_at_position(x, y, length, width)
        
        if building.can_be_boosted():
            if culture_recue >= building.boost_100:
                boost = 2.0
                boost_level = "🔥 100%"
            elif culture_recue >= building.boost_50:
                boost = 1.5
                boost_level = "✨ 50%"
            elif culture_recue >= building.boost_25:
                boost = 1.25
                boost_level = "⭐ 25%"
            else:
                boost = 1.0
                boost_level = "⚪ 0%"
        else:
            boost = 1.0
            boost_level = "⚪ Producteur"
        
        placed_info = {
            'building': building,
            'x': x, 'y': y,
            'orientation': orientation,
            'length': length,
            'width': width,
            'building_id': building_id,
            'culture_recue': culture_recue,
            'boost': boost,
            'boost_level': boost_level
        }
        self.placed_buildings.append(placed_info)
        
        if building.is_producer():
            center_x = x + length // 2
            center_y = y + width // 2
            self.producer_positions.append((center_x, center_y, building, building_id, building.radius))
        else:
            center_x = x + length // 2
            center_y = y + width // 2
            self.consumer_positions.append((center_x, center_y, building, building_id))
        
        return True
    
    def place_all_buildings(self) -> dict:
        self.initialize_grids()
        self.placed_buildings = []
        self.producer_positions = []
        self.consumer_positions = []
        
        total_to_place = sum(b.quantity for b in self.buildings)
        
        # Séparer producteurs et consommateurs
        producers = []
        consumers = []
        for building in self.buildings:
            for _ in range(building.quantity):
                if building.is_producer():
                    producers.append(building)
                else:
                    consumers.append(building)
        
        st.info(f"🎯 {len(producers)} producteurs, {len(consumers)} consommateurs")
        
        # Créer l'optimiseur de boosts
        optimizer = BoostOptimizer(producers, consumers)
        clusters = optimizer.find_best_clusters()
        st.info(f"🔗 {len(clusters)} clusters de boost identifiés")
        
        # Identifier les zones
        self.find_available_zones()
        self.available_zones.sort(key=len, reverse=True)
        
        # PHASE 1: Placer les clusters
        st.info("📌 Phase 1: Placement des clusters de boost...")
        
        clusters_placed = 0
        for cluster in clusters:
            if not self.available_zones:
                break
            
            # Prendre la plus grande zone disponible
            zone = self.available_zones[0]
            
            if self.place_cluster(cluster, zone):
                clusters_placed += 1
                
                # Mettre à jour la zone
                new_zone = []
                for (zx, zy) in zone:
                    if self.placement_grid[zx, zy] == 0:
                        new_zone.append((zx, zy))
                
                if new_zone:
                    self.available_zones[0] = new_zone
                else:
                    self.available_zones.pop(0)
                
                self.available_zones.sort(key=len, reverse=True)
        
        st.info(f"✅ {clusters_placed} clusters placés")
        
        # PHASE 2: Placer les producteurs restants
        remaining_producers = [b for b in producers if b not in [c[0] for c in clusters]]
        
        for producer in remaining_producers:
            if not self.available_zones:
                break
            
            zone = self.available_zones[0]
            xs = [p[0] for p in zone]
            ys = [p[1] for p in zone]
            min_x, max_x = min(xs), max(xs)
            min_y, max_y = min(ys), max(ys)
            center_x = (min_x + max_x) // 2
            center_y = (min_y + max_y) // 2
            
            placed = False
            for orientation in ['H', 'V']:
                length, width = producer.get_dimensions(orientation)
                
                for dx in [-2, -1, 0, 1, 2]:
                    for dy in [-2, -1, 0, 1, 2]:
                        x = max(min_x, min(center_x + dx, max_x - length + 1))
                        y = max(min_y, min(center_y + dy, max_y - width + 1))
                        
                        if self.can_place(x, y, length, width):
                            if self.place_building(producer, x, y, orientation):
                                placed = True
                                break
                    if placed:
                        break
                if placed:
                    break
            
            if placed:
                new_zone = []
                for (zx, zy) in zone:
                    if self.placement_grid[zx, zy] == 0:
                        new_zone.append((zx, zy))
                
                if new_zone:
                    self.available_zones[0] = new_zone
                else:
                    self.available_zones.pop(0)
                
                self.available_zones.sort(key=len, reverse=True)
        
        # PHASE 3: Placer les consommateurs restants dans les rayons
        remaining_consumers = [c for c in consumers if c not in sum([c[1:] for c in clusters], [])]
        
        for consumer in remaining_consumers:
            best_placement = None
            best_culture = -1
            best_zone_idx = -1
            
            for zone_idx, zone in enumerate(self.available_zones):
                xs = [p[0] for p in zone]
                ys = [p[1] for p in zone]
                min_x, max_x = min(xs), max(xs)
                min_y, max_y = min(ys), max(ys)
                
                for orientation in ['H', 'V']:
                    length, width = consumer.get_dimensions(orientation)
                    
                    if length > (max_x - min_x + 1) or width > (max_y - min_y + 1):
                        continue
                    
                    for x in range(min_x, max_x - length + 2):
                        for y in range(min_y, max_y - width + 2):
                            if self.can_place(x, y, length, width):
                                culture = self.calculate_culture_at_position(x, y, length, width)
                                if culture > best_culture:
                                    best_culture = culture
                                    best_placement = (x, y, orientation)
                                    best_zone_idx = zone_idx
            
            if best_placement and best_culture > 0:
                x, y, orientation = best_placement
                self.place_building(consumer, x, y, orientation)
                
                if best_zone_idx >= 0:
                    zone = self.available_zones[best_zone_idx]
                    new_zone = []
                    for (zx, zy) in zone:
                        if self.placement_grid[zx, zy] == 0:
                            new_zone.append((zx, zy))
                    
                    if new_zone:
                        self.available_zones[best_zone_idx] = new_zone
                    else:
                        self.available_zones.pop(best_zone_idx)
        
        # PHASE 4: Placer le reste (ceux qui n'ont pas pu être boostés)
        remaining = remaining_producers + remaining_consumers
        for building in remaining:
            placed = False
            for zone_idx, zone in enumerate(self.available_zones):
                if placed:
                    break
                
                xs = [p[0] for p in zone]
                ys = [p[1] for p in zone]
                min_x, max_x = min(xs), max(xs)
                min_y, max_y = min(ys), max(ys)
                
                for orientation in ['H', 'V']:
                    if placed:
                        break
                    length, width = building.get_dimensions(orientation)
                    
                    if length > (max_x - min_x + 1) or width > (max_y - min_y + 1):
                        continue
                    
                    for x in range(min_x, max_x - length + 2):
                        if placed:
                            break
                        for y in range(min_y, max_y - width + 2):
                            if self.can_place(x, y, length, width):
                                self.place_building(building, x, y, orientation)
                                placed = True
                                break
        
        # Statistiques finales
        total_culture_produced = sum(p['building'].culture_produced for p in self.placed_buildings if p['building'].is_producer())
        
        boost_stats = {'100%': 0, '50%': 0, '25%': 0, '0%': 0}
        boost_details = []
        
        for p in self.placed_buildings:
            if p['building'].can_be_boosted():
                if p['culture_recue'] >= p['building'].boost_100:
                    boost_stats['100%'] += 1
                    boost_details.append(f"{p['building'].name}: {p['culture_recue']:.0f} culture (100%)")
                elif p['culture_recue'] >= p['building'].boost_50:
                    boost_stats['50%'] += 1
                    boost_details.append(f"{p['building'].name}: {p['culture_recue']:.0f} culture (50%)")
                elif p['culture_recue'] >= p['building'].boost_25:
                    boost_stats['25%'] += 1
                    boost_details.append(f"{p['building'].name}: {p['culture_recue']:.0f} culture (25%)")
                else:
                    boost_stats['0%'] += 1
                    if p['culture_recue'] > 0:
                        boost_details.append(f"{p['building'].name}: {p['culture_recue']:.0f} culture (insuffisant)")
        
        cultures_recues = [p['culture_recue'] for p in self.placed_buildings if p['building'].can_be_boosted()]
        culture_moyenne = sum(cultures_recues) / len(cultures_recues) if cultures_recues else 0
        consumers_with_culture = sum(1 for p in self.placed_buildings if p['building'].can_be_boosted() and p['culture_recue'] > 0)
        total_consumers = sum(1 for p in self.placed_buildings if p['building'].can_be_boosted())
        
        return {
            'total_buildings': len(self.placed_buildings),
            'buildings_total': total_to_place,
            'total_culture_produced': total_culture_produced,
            'culture_moyenne_recue': culture_moyenne,
            'consumers_with_culture': consumers_with_culture,
            'total_consumers': total_consumers,
            'grid': self.placement_grid.copy(),
            'placed_buildings': copy.deepcopy(self.placed_buildings),
            'boost_stats': boost_stats,
            'boost_details': boost_details,
            'producers_placed': len([p for p in self.placed_buildings if p['building'].is_producer()]),
            'consumers_placed': total_consumers,
            'remaining': len(remaining)
        }

def create_visual_excel_sheet(writer, grid: np.ndarray, placed_buildings: List[dict]):
    """Crée un onglet Excel avec une visualisation couleur des bâtiments"""
    
    viz_data = []
    for i in range(grid.shape[0]):
        row = []
        for j in range(grid.shape[1]):
            val = grid[i, j]
            if val == -1:
                row.append("█")
            elif val == 0:
                row.append("·")
            else:
                building = next((p for p in placed_buildings if p['building_id'] == val), None)
                if building:
                    row.append(building['building'].name[:3])
                else:
                    row.append(str(val))
        viz_data.append(row)
    
    df_viz = pd.DataFrame(viz_data)
    df_viz.to_excel(writer, sheet_name='Carte_couleurs', index=False, header=False)
    
    workbook = writer.book
    worksheet = writer.sheets['Carte_couleurs']
    
    colors = [
        'FF6B6B', '4ECDC4', '45B7D1', '96CEB4', 'FFEEAD', 'D4A5A5',
        '9B59B6', '3498DB', 'E67E22', '2ECC71', 'E74C3C', '1ABC9C',
        'F1C40F', 'E67E22', '9B59B6', '34495E', '16A085', '27AE60'
    ]
    
    building_colors = {}
    for i, placed in enumerate(placed_buildings):
        building_id = placed['building_id']
        building_colors[building_id] = colors[i % len(colors)]
    
    for i in range(grid.shape[0]):
        for j in range(grid.shape[1]):
            cell = worksheet.cell(row=i+1, column=j+1)
            val = grid[i, j]
            
            if val == -1:
                cell.fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
                cell.font = Font(color='FFFFFF', size=8)
            elif val == 0:
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                cell.font = Font(color='000000', size=8)
            else:
                color = building_colors.get(val, 'CCCCCC')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                cell.font = Font(color='000000', size=8, bold=True)
            
            cell.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.column_dimensions[get_column_letter(j+1)].width = 3

def create_building_summary(placed_buildings: List[dict]) -> pd.DataFrame:
    """Crée un résumé de TOUS les bâtiments placés"""
    summary = []
    for p in placed_buildings:
        if p['building'].can_be_boosted():
            seuils = f"{p['building'].boost_25:.0f}/{p['building'].boost_50:.0f}/{p['building'].boost_100:.0f}"
        else:
            seuils = "-"
        
        summary.append({
            'ID': p['building_id'],
            'Bâtiment': p['building'].name,
            'Type': '🏭 Producteur' if p['building'].is_producer() else '🏠 Consommateur',
            'X': p['x'],
            'Y': p['y'],
            'Dimensions': f"{p['length']}x{p['width']}",
            'Culture prod': p['building'].culture_produced if p['building'].is_producer() else 0,
            'Culture reçue': f"{p['culture_recue']:.0f}",
            'Boost': p['boost_level'],
            'Seuils': seuils,
            'Rayon': p['building'].radius if p['building'].is_producer() else '-'
        })
    return pd.DataFrame(summary)

# Interface Streamlit
st.set_page_config(page_title="Optimiseur de Bâtiments", page_icon="🏗️", layout="wide")

st.title("🏗️ Optimiseur de Placement de Bâtiments - Version Boost Maximisé")
st.markdown("""
**Objectif :** Maximiser les boosts en créant des clusters producteurs + consommateurs

**Stratégie :**
1. **Phase 1** : Identification des clusters de boost optimaux
2. **Phase 2** : Placement des clusters dans les zones libres
3. **Phase 3** : Placement des consommateurs restants dans les rayons
4. **Phase 4** : Placement final des bâtiments non boostés
""")

with st.sidebar:
    st.header("📁 Chargement")
    uploaded_file = st.file_uploader("Fichier Excel", type=['xlsx', 'xls'])
    optimize_button = st.button("🚀 Lancer l'optimisation boost maximale", type="primary")

if uploaded_file and optimize_button:
    try:
        # Lecture
        df_terrain = pd.read_excel(uploaded_file, sheet_name=0, header=None, engine='openpyxl')
        df_buildings = pd.read_excel(uploaded_file, sheet_name=1, engine='openpyxl')
        
        terrain_grid = df_terrain.values.astype(int)
        
        # Création des bâtiments
        buildings = []
        for _, row in df_buildings.iterrows():
            def get_float(col, default=0):
                val = row[col] if col in row else default
                return float(val) if not pd.isna(val) else default
            
            def get_int(col, default=1):
                val = row[col] if col in row else default
                return int(val) if not pd.isna(val) else default
            
            building = Building(
                name=str(row['Nom']),
                length=get_int('longueur', 1),
                width=get_int('largeur', 1),
                quantity=get_int('quantité', 1),
                culture_produced=get_float('culture', 0),
                radius=get_int('rayonnement', 0),
                boost_25=get_float('Boost 25%', 0),
                boost_50=get_float('Boost 50%', 0),
                boost_100=get_float('Boost 100%', 0)
            )
            buildings.append(building)
        
        total_demande = sum(b.quantity for b in buildings)
        cells_libres = np.sum(terrain_grid == 1)
        
        n_producers = sum(b.quantity for b in buildings if b.is_producer())
        n_consumers = sum(b.quantity for b in buildings if b.can_be_boosted())
        
        st.success(f"✅ {len(buildings)} types, {total_demande} bâtiments")
        st.info(f"🗺️ Terrain: {cells_libres} cases libres")
        st.info(f"🎯 {n_producers} producteurs, {n_consumers} consommateurs")
        
        # Placement
        placer = BuildingPlacer(terrain_grid, buildings)
        results = placer.place_all_buildings()
        
        # Résultats
        st.subheader("📊 Résultats")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("🏢 Placés", f"{results['total_buildings']}/{total_demande}")
        with col2:
            st.metric("💰 Culture produite", f"{results['total_culture_produced']:.0f}")
        with col3:
            st.metric("📊 Dans un rayon", f"{results['consumers_with_culture']}/{results['total_consumers']}")
        with col4:
            st.metric("🎯 Culture moyenne", f"{results['culture_moyenne_recue']:.0f}")
        
        # Statistiques des boosts
        st.subheader("⚡ Distribution des boosts")
        boost_stats = results['boost_stats']
        
        cols = st.columns(4)
        colors = ['gold', 'lightgreen', 'lightblue', 'lightgray']
        for i, (label, count) in enumerate(boost_stats.items()):
            with cols[i]:
                st.markdown(f"""
                <div style='background-color: {colors[i]}; padding: 10px; border-radius: 5px; text-align: center'>
                    <h3 style='margin:0'>{label}</h3>
                    <h2 style='margin:0'>{count}</h2>
                </div>
                """, unsafe_allow_html=True)
        
        # Détail des boosts
        if results['boost_details']:
            with st.expander("📋 Détail des boosts par bâtiment"):
                for detail in results['boost_details'][:20]:  # Limiter à 20 pour lisibilité
                    st.write(detail)
        
        # Visualisation
        st.subheader("🗺️ Carte de placement")
        fig = go.Figure(data=go.Heatmap(
            z=results['grid'],
            colorscale='Viridis',
            showscale=False,
            text=results['grid'],
            texttemplate="%{text}",
            textfont={"size": 6}
        ))
        fig.update_layout(height=600, yaxis=dict(autorange='reversed'))
        st.plotly_chart(fig, use_container_width=True)
        
        # Tableau récapitulatif
        st.subheader("📋 Liste de tous les bâtiments placés")
        df_summary = create_building_summary(results['placed_buildings'])
        st.dataframe(df_summary, use_container_width=True)
        
        # Export
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Tous_les_batiments', index=False)
            create_visual_excel_sheet(writer, results['grid'], results['placed_buildings'])
            
            stats = {
                'Total bâtiments': results['total_buildings'],
                'Producteurs placés': results['producers_placed'],
                'Consommateurs placés': results['consumers_placed'],
                'Culture produite': results['total_culture_produced'],
                'Culture moyenne reçue': f"{results['culture_moyenne_recue']:.0f}",
                'Boost 100%': boost_stats['100%'],
                'Boost 50%': boost_stats['50%'],
                'Boost 25%': boost_stats['25%'],
                'Sans boost': boost_stats['0%'],
                'Non placés': total_demande - results['total_buildings']
            }
            pd.DataFrame([stats]).to_excel(writer, sheet_name='Statistiques', index=False)
        
        st.download_button(
            label="📥 Télécharger le fichier Excel",
            data=output.getvalue(),
            file_name="placement_boosts_maximises.xlsx"
        )
        
    except Exception as e:
        st.error(f"❌ Erreur: {str(e)}")
        st.exception(e)